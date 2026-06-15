import asyncio
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services import sanity_service


async def main():
    try:
        # 1. Fetch data
        query = '*[_type == "product" && !defined(image)]{ name, brandName, genericName }'
        products = await sanity_service.query_sanity(query)

        # 2. Setup workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Images"

        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True

        # 3. Styling definitions
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10)
        fill_header = PatternFill(start_color="1D9FDA",
                                  end_color="1D9FDA",
                                  fill_type="solid")  # Getmeds brand blue
        fill_zebra = PatternFill(start_color="F7FAFC",
                                 end_color="F7FAFC",
                                 fill_type="solid")  # light gray/blue

        border_thin = Side(border_style="thin", color="E2E8F0")
        border_all = Border(left=border_thin,
                            right=border_thin,
                            top=border_thin,
                            bottom=border_thin)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # 4. Write Headers
        headers = ["#", "Product Name", "Brand Name", "Generic Name"]
        ws.append(headers)

        # Format Headers
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if col_idx == 1 else align_left
            cell.border = border_all

        # 5. Write Data Rows
        for idx, p in enumerate(products, 1):
            row_idx = idx + 1
            ws.row_dimensions[row_idx].height = 20

            row_data = [
                idx,
                p.get("name") or "N/A",
                p.get("brandName") or "N/A",
                p.get("genericName") or "N/A"
            ]
            ws.append(row_data)

            # Apply styling to data row cells
            is_zebra = (idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = border_all
                cell.alignment = align_center if col_idx == 1 else align_left
                if is_zebra:
                    cell.fill = fill_zebra

        # 6. Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # 7. Save to database root
        output_dir = r"c:\Users\Getmeds\Desktop\getmeds_database"
        output_filename = "products_missing_images.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        wb.save(output_path)
        print(f"SUCCESS: Created excel file at {output_path}")

    except Exception as e:
        print("Error during Excel export:", e)


if __name__ == "__main__":
    asyncio.run(main())
