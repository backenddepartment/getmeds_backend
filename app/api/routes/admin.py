# pyrefly: ignore [missing-import]
from fastapi import APIRouter, HTTPException, Depends, Header, Request, Response, UploadFile, File, Form
# pyrefly: ignore [missing-import]
from pydantic import BaseModel
from app.services.sanity_service import sanity_service
from app.services.security_service import security_service
from app.core.config import get_settings
from app.core.schema_loader import schema_loader
import uuid
import re
import base64
import time

router = APIRouter()
settings = get_settings()

class LoginRequest(BaseModel):
    email: str
    password: str

def generate_access_token(username: str, role: str, permissions: str):
    expiry = int(time.time()) + 900  # 15 minutes
    raw_str = f"{username}:{role}:{permissions}:{expiry}"
    encoded = base64.b64encode(raw_str.encode("utf-8")).decode("utf-8")
    return f"GMP-SEC-ACCESS-{encoded}"

def decode_access_token(token: str):
    if not token.startswith("GMP-SEC-ACCESS-"):
        return None
    try:
        encoded_part = token[len("GMP-SEC-ACCESS-"):]
        decoded_str = base64.b64decode(encoded_part.encode("utf-8")).decode("utf-8")
        parts = decoded_str.split(":", 3)
        if len(parts) == 4:
            username, role, permissions, expiry = parts
            if int(time.time()) < int(expiry):
                return {"username": username, "role": role, "permissions": permissions}
    except Exception:
        pass
    return None

def generate_refresh_token(username: str, passcode: str):
    expiry = int(time.time()) + 604800  # 7 days
    raw_str = f"{username}:{passcode}:{expiry}"
    encoded = base64.b64encode(raw_str.encode("utf-8")).decode("utf-8")
    return f"GMP-SEC-REFRESH-{encoded}"

def decode_refresh_token(token: str):
    if not token.startswith("GMP-SEC-REFRESH-"):
        return None
    try:
        encoded_part = token[len("GMP-SEC-REFRESH-"):]
        decoded_str = base64.b64decode(encoded_part.encode("utf-8")).decode("utf-8")
        parts = decoded_str.split(":", 2)
        if len(parts) == 3:
            username, passcode, expiry = parts
            if int(time.time()) < int(expiry):
                return {"username": username, "passcode": passcode}
    except Exception:
        pass
    return None

async def verify_admin_token(request: Request, x_admin_token: str = Header(None, alias="X-Admin-Token")):
    expected_token = getattr(settings, "ADMIN_TOKEN", "getmeds-admin-secret-key")
    
    # 1. Fallback for the master system secret token
    if x_admin_token == expected_token:
        return
        
    # 2. Extract access token from header or fallback to cookie
    token = x_admin_token or request.cookies.get("gmp_access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Missing admin session token.")
        
    claims = decode_access_token(token)
    if not claims:
        raise HTTPException(status_code=401, detail="Expired or invalid admin session token.")
        
    # Query Supabase PostgreSQL to ensure operator is Active
    username = claims["username"]
    try:
        user = await security_service.get_user_by_username(username)
    except Exception as e:
        print(f"WARNING: Supabase unreachable for token validation: {e}. Allowing access for {username}.")
        return
    
    if not user:
        raise HTTPException(status_code=401, detail="Operator session is invalid.")
        
    if user.get("status") != "Active":
        raise HTTPException(status_code=403, detail="Operator account status is suspended.")

@router.post("/auth/login")
async def admin_auth_login(payload: LoginRequest, response: Response):
    """
    Authenticates operator by email and password.
    Stores session tokens securely in cookies and returns claims for client UI state.
    """
    res = await security_service.authenticate_admin(email=payload.email, password=payload.password)
    if not res.get("authenticated"):
        raise HTTPException(status_code=401, detail=res.get("detail", "Authentication failed."))
        
    username = res["username"]
    role = res["role"]
    permissions = res["permissions"]
    
    # Generate short-lived Access and long-lived Refresh tokens
    access_token = generate_access_token(username, role, permissions)
    refresh_token = generate_refresh_token(username, payload.password)
    
    # Register secure HTTP-only cookies
    response.set_cookie(
        key="gmp_access_token",
        value=access_token,
        max_age=900,  # 15 mins
        httponly=True,
        secure=False,  # localhost compatibility
        samesite="lax"
    )
    response.set_cookie(
        key="gmp_refresh_token",
        value=refresh_token,
        max_age=604800,  # 7 days
        httponly=True,
        secure=False,
        samesite="lax"
    )
    
    # Retrieve RBAC permissions to determine allowed collections
    matrix = await security_service.get_role_permissions()
    # v2 format: { role: { collection: [actions] } }
    # v1 legacy: { collection: [roles] }
    is_v2 = matrix and not any(isinstance(v, list) for v in matrix.values())
    if is_v2:
        role_perms = matrix.get(role, {})
        allowed = [col for col, actions in role_perms.items() if 'read' in actions]
    else:
        allowed = [col for col, roles in matrix.items() if role in roles]
        role_perms = {col: (['read', 'create', 'edit', 'delete'] if role in roles else ['read']) for col, roles in matrix.items()}

    return {
        "status": "success",
        "username": username,
        "role": role,
        "permissions": permissions,
        "allowedCollections": allowed,
        "roleActionPerms": role_perms,
        "token": access_token
    }

class GoogleSSORequest(BaseModel):
    credential: str

@router.post("/auth/google-sso")
async def google_sso_auth(payload: GoogleSSORequest, response: Response):
    """
    Validates Google SSO JWT, checks permissions on Supabase, 
    and issues secure admin session cookies.
    """
    # pyrefly: ignore [missing-import]
    import httpx
    try:
        token = payload.credential
        async with httpx.AsyncClient(timeout=10.0) as client:
            google_res = await client.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={token}")
            
        if google_res.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid Google SSO token.")
            
        google_data = google_res.json()
        email = google_data.get("email")
        if not email:
            raise HTTPException(status_code=400, detail="Google token does not contain email.")
            
        user = await security_service.get_user_by_username(email)
        if not user:
            await security_service.add_security_log(
                event=f"Failed login attempt: Unauthorized Google email {email}",
                user="Unknown",
                ip_address="127.0.0.1",
                status="FAILED",
                severity="HIGH"
            )
            raise HTTPException(
                status_code=403, 
                detail=f"Access Denied: Your Google account ({email}) is not authorized."
            )
            
        if user.get("status") != "Active":
            await security_service.add_security_log(
                event=f"Blocked login attempt for suspended operator email: {email}",
                user=user["username"],
                ip_address="127.0.0.1",
                status="FAILED",
                severity="WARNING"
            )
            raise HTTPException(
                status_code=403, 
                detail="Access Blocked: Your administrative account is currently suspended."
            )
            
        username = user["username"]
        role = user["role"]
        permissions = user["permissions"]
        passcode = user["passcode"]
        
        await security_service.add_security_log(
            event=f"Google SSO authentication successful for {username}",
            user=username,
            ip_address="127.0.0.1",
            status="SUCCESS",
            severity="INFO"
        )
        
        access_token = generate_access_token(username, role, permissions)
        refresh_token = generate_refresh_token(username, passcode)
        
        response.set_cookie(key="gmp_access_token", value=access_token, max_age=900, httponly=True, secure=False, samesite="lax")
        response.set_cookie(key="gmp_refresh_token", value=refresh_token, max_age=604800, httponly=True, secure=False, samesite="lax")
        
        # Retrieve RBAC permissions to determine allowed collections
        matrix = await security_service.get_role_permissions()
        is_v2 = matrix and not any(isinstance(v, list) for v in matrix.values())
        if is_v2:
            role_perms = matrix.get(role, {})
            allowed = [col for col, actions in role_perms.items() if 'read' in actions]
        else:
            allowed = [col for col, roles in matrix.items() if role in roles]
            role_perms = {col: (['read','create','edit','delete'] if role in roles else ['read']) for col, roles in matrix.items()}
            
        return {
            "status": "success",
            "username": username,
            "role": role,
            "permissions": permissions,
            "allowedCollections": allowed,
            "roleActionPerms": role_perms,
            "token": access_token
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SSO verification failed: {str(e)}")

@router.post("/auth/refresh")
async def admin_auth_refresh(request: Request, response: Response):
    """
    Seamless access token rotation using HTTP-only Refresh Token.
    """
    refresh_token = request.cookies.get("gmp_refresh_token")
    if not refresh_token:
        raise HTTPException(status_code=401, detail="Missing refresh token.")
        
    claims = decode_refresh_token(refresh_token)
    if not claims:
        response.delete_cookie("gmp_access_token")
        response.delete_cookie("gmp_refresh_token")
        raise HTTPException(status_code=401, detail="Expired or invalid refresh token.")
        
    username = claims["username"]
    passcode = claims["passcode"]
    
    try:
        user = await security_service.get_user_by_username(username)
        if not user or user.get("passcode") != passcode or user.get("status") != "Active":
            response.delete_cookie("gmp_access_token")
            response.delete_cookie("gmp_refresh_token")
            raise HTTPException(status_code=403, detail="Operator access suspended or modified.")
        role = user["role"]
        permissions = user["permissions"]
    except HTTPException:
        raise
    except Exception as db_err:
        import sys
        print(f"WARNING: Supabase database is unreachable during admin_auth_refresh: {db_err}. Falling back to in-memory validation.", file=sys.stderr)
        # Fallback credentials seeder match
        fallbacks = [
            {
                "username": "admin",
                "role": "Administrator",
                "permissions": "full_access",
                "passcode": "Getmeds@1"
            },
            {
                "username": "backenddepartment@gmail.com",
                "role": "Administrator",
                "permissions": "Full System Access, CMS Mutations, Schema Control",
                "passcode": "GMP-ADMIN-2026"
            },
            {
                "username": "Getmeds Admin",
                "role": "Administrator",
                "permissions": "Full System Access, CMS Mutations, Schema Control",
                "passcode": "GMP-GETMEDS-2026"
            }
        ]
        matched_fb = None
        for fb_user in fallbacks:
            if fb_user["username"] == username and fb_user["passcode"] == passcode:
                matched_fb = fb_user
                break
        if not matched_fb:
            response.delete_cookie("gmp_access_token")
            response.delete_cookie("gmp_refresh_token")
            raise HTTPException(status_code=403, detail="Operator access suspended or invalid.")
        user = matched_fb
        role = user["role"]
        permissions = user["permissions"]

    new_access = generate_access_token(username, role, permissions)
    new_refresh = generate_refresh_token(username, passcode)
    
    response.set_cookie(key="gmp_access_token", value=new_access, max_age=900, httponly=True, secure=False, samesite="lax")
    response.set_cookie(key="gmp_refresh_token", value=new_refresh, max_age=604800, httponly=True, secure=False, samesite="lax")
    
    # Retrieve RBAC permissions to determine allowed collections
    try:
        matrix = await security_service.get_role_permissions()
    except Exception:
        matrix = None

    is_v2 = matrix and not any(isinstance(v, list) for v in matrix.values())
    if is_v2:
        role_perms = matrix.get(role, {})
        allowed = [col for col, actions in role_perms.items() if 'read' in actions]
    else:
        # Default or fallback role-matrix configuration
        if role == "Administrator":
            allowed = ["product", "category", "heroSlide", "service", "dealOfDay", "categoryBanner", "team", "faq", "chatSession", "_security"]
            role_perms = {col: ['read', 'create', 'edit', 'delete'] for col in allowed}
        else:
            allowed = ["product", "category", "heroSlide", "service", "dealOfDay", "categoryBanner", "team", "faq", "chatSession"]
            role_perms = {col: ['read', 'create', 'edit'] for col in allowed}

    return {
        "status": "success",
        "token": new_access,
        "allowedCollections": allowed,
        "roleActionPerms": role_perms
    }

@router.post("/auth/logout")
async def admin_auth_logout(response: Response):
    """
    Purges secure HTTP-only cookies and ends active administrative session.
    """
    response.delete_cookie("gmp_access_token")
    response.delete_cookie("gmp_refresh_token")
    return {"status": "success"}

class DynamicListProxy(list):
    def __contains__(self, item):
        return item in schema_loader.get_supported_types()
    def __iter__(self):
        return iter(schema_loader.get_supported_types())

class DynamicDictProxy(dict):
    def __init__(self, fetcher):
        self.fetcher = fetcher
    def get(self, key, default=None):
        return self.fetcher().get(key, default)
    def __getitem__(self, key):
        return self.fetcher()[key]
    def __contains__(self, key):
        return key in self.fetcher()
    def keys(self):
        return self.fetcher().keys()
    def values(self):
        return self.fetcher().values()
    def items(self):
        return self.fetcher().items()

SUPPORTED_TYPES = DynamicListProxy()


@router.get("/schemas", dependencies=[Depends(verify_admin_token)])
async def get_schemas():
    """
    Returns the complete dynamic schemas, including field configurations,
    table column views, and import specs.
    """
    try:
        from app.core.config import get_settings
        _settings = get_settings()
        return {
            "status": "success",
            "field_schemas": schema_loader.get_field_schemas(),
            "table_schemas": schema_loader.get_table_schemas(),
            "import_field_schemas": schema_loader.get_import_field_schemas(),
            "collection_meta": schema_loader.get_collection_meta(),
            "image_collections": schema_loader.get_image_collections(),
            "sanity_project_id": _settings.SANITY_PROJECT_ID,
            "sanity_dataset": _settings.SANITY_DATASET
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/documents", dependencies=[Depends(verify_admin_token)])
async def get_documents(collection: str = "product", page: int = 1, limit: int = 5):
    """
    Returns a server-side paginated list of existing documents for a specific active collection schema.
    """
    if collection not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported collection type: {collection}")
        
    try:
        # 1. Fetch only the active collection's total count
        count_query = f'count(*[_type == "{collection}"])'
        total_count = await sanity_service.query_sanity(count_query)
        
        # 2. Query target paginated slice using type-specific projection
        start = (page - 1) * limit
        end = page * limit
        
        projection = GROQ_PROJECTIONS.get(collection, '{ _id, _type, _createdAt }')
        groq_query = f'*[_type == "{collection}"] {projection} | order(_createdAt desc) [{start}...{end}]'
        results = await sanity_service.query_sanity(groq_query)
        
        return {
            "status": "success",
            "documents": results,
            "total": total_count,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

PREFIX_MAP = DynamicDictProxy(schema_loader.get_prefix_map)
GROQ_PROJECTIONS = DynamicDictProxy(schema_loader.get_groq_projections)

@router.post("/collection/{doc_type}", dependencies=[Depends(verify_admin_token)])
async def create_collection_document(doc_type: str, data: dict):
    """
    Unified dynamic creator endpoint. Accepts custom payloads for ANY active database collection schema,
    generates clean custom sequential prefix IDs (e.g. GMP0000-0001), performs auto-slugification,
    and executes Sanity mutations cleanly.
    """
    if doc_type not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported document type: {doc_type}")
    
    try:
        # Query existing count in database to form next sequential custom ID (checking both draft and published counts)
        count_query = f'count(*[_type == "{doc_type}" && !(_id in path("drafts.**"))])'
        count = await sanity_service.query_sanity(count_query)
        new_idx = count + 1
        prefix = PREFIX_MAP.get(doc_type, f"{doc_type}0000-")
        doc_id = f"{prefix}{new_idx:04d}"
        draft_id = f"drafts.{doc_id}"
        
        # Build base document mutation payload
        doc_payload = {
            "_id": draft_id,
            "_type": doc_type,
            **data
        }

        # Auto slugify if slug field is present in schema
        slug_obj = _auto_slugify(doc_type, data)
        if slug_obj:
            doc_payload["slug"] = slug_obj

        mutations = [{"create": doc_payload}]
        result = await sanity_service.mutate_sanity(mutations)
        return {"status": "success", "id": draft_id, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/document/{doc_id}", dependencies=[Depends(verify_admin_token)])
async def get_document_detail(doc_id: str):
    """
    Fetches the complete raw document by _id.
    - Supabase UUIDs (admin_users, access_points, security_logs) → resolved via security_service
    - Legacy-prefixed security IDs (GMAU/GMAP/GMSL) → resolved via security_service  
    - All other Sanity CMS document IDs → resolved via Sanity GROQ
    """
    import re
    _UUID_RE = re.compile(
        r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        re.IGNORECASE
    )

    try:
        is_security_record = False
        result = None
        if _UUID_RE.match(doc_id) or doc_id.isdigit() or doc_id.startswith(("GMAU", "GMAP", "GMSL")):
            # Try to fetch from Supabase
            result = await security_service.get_security_record_by_id(doc_id)
            if result:
                is_security_record = True
                # Map table name to frontend UI _type for FIELD_SCHEMAS lookup
                if result.get("_table") == "admin_users":
                    result["_type"] = "adminUser"
                elif result.get("_table") == "access_points":
                    result["_type"] = "accessPoint"
            elif doc_id.startswith(("GMAU", "GMAP", "GMSL")):
                raise HTTPException(status_code=404, detail=f"Security record '{doc_id}' not found in Supabase.")
        
        if not is_security_record:
            groq_query = f'*[_id in ["{doc_id}", "drafts.{doc_id}"]]'
            docs = await sanity_service.query_sanity(groq_query) or []
            
            draft_doc = next((d for d in docs if d["_id"].startswith("drafts.")), None)
            pub_doc = next((d for d in docs if not d["_id"].startswith("drafts.")), None)
            
            if not draft_doc and not pub_doc:
                raise HTTPException(status_code=404, detail=f"Document '{doc_id}' not found.")
            
            result = draft_doc if draft_doc else pub_doc
            result["_hasDraft"] = draft_doc is not None
            result["_isPublished"] = pub_doc is not None
            result["_status"] = "draft" if (draft_doc and not pub_doc) else ("modified" if (draft_doc and pub_doc) else "published")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

    return {"status": "success", "document": result}

@router.patch("/document/{doc_id}", dependencies=[Depends(verify_admin_token)])
async def update_document(doc_id: str, data: dict):
    """
    Patches specific fields of an existing document.
    Security-prefixed IDs or Supabase UUIDs update the Supabase admin_users or access_points table.
    All other IDs patch Sanity CMS documents.
    """
    if not data:
        raise HTTPException(status_code=400, detail="No fields provided to update.")
    
    import re
    _UUID_RE = re.compile(
        r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        re.IGNORECASE
    )

    try:
        is_security_record = False
        record = None
        if _UUID_RE.match(doc_id) or doc_id.isdigit() or doc_id.startswith(("GMAU", "GMAP", "GMSL")):
            record = await security_service.get_security_record_by_id(doc_id)
            if record:
                is_security_record = True
            elif doc_id.startswith(("GMAU", "GMAP", "GMSL")):
                raise HTTPException(status_code=404, detail="Security record not found in Supabase.")

        if is_security_record:
            table = record.get("_table")
            if table == "admin_users":
                await security_service.update_admin_user(doc_id, data)
            elif table == "access_points":
                await security_service.update_access_point(doc_id, data)
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported security record table update: {table}")
            return {"status": "success", "id": doc_id}
        else:
            base_id = doc_id[7:] if doc_id.startswith("drafts.") else doc_id
            draft_id = f"drafts.{base_id}"
            
            # Check if draft exists
            draft_exists = await sanity_service.query_sanity(f'count(*[_id == "{draft_id}"])')
            if not draft_exists:
                # Fetch published doc to copy it to draft
                pub_doc = await sanity_service.query_sanity(f'*[_id == "{base_id}"][0]')
                if pub_doc:
                    # Strip system keys
                    for key in ("_createdAt", "_updatedAt", "_rev"):
                        pub_doc.pop(key, None)
                    pub_doc["_id"] = draft_id
                    # Create the draft doc copy
                    await sanity_service.mutate_sanity([{"create": pub_doc}])
            
            for key in ("_id", "_type", "_rev", "_createdAt", "_updatedAt"):
                data.pop(key, None)
            
            mutations = [{"patch": {"id": draft_id, "set": data}}]
            result = await sanity_service.mutate_sanity(mutations)
            return {"status": "success", "id": draft_id, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/assets/upload", dependencies=[Depends(verify_admin_token)])
async def upload_asset_generic(request: Request):
    """
    Uploads a raw image file directly to Sanity assets and returns the asset reference.
    Expects raw binary body with Content-Type: image/*, e.g. image/jpeg or image/png.
    """
    from app.core.config import get_settings
    _settings = get_settings()
    content_type = request.headers.get("content-type", "image/jpeg")
    image_bytes = await request.body()

    if not image_bytes:
        raise HTTPException(status_code=400, detail="Empty image body.")

    # Determine file extension from content-type
    ext_map = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}
    ext = ext_map.get(content_type.split(";")[0].strip(), "jpg")

    try:
        # pyrefly: ignore [missing-import]
        import httpx
        asset_url = (
            f"https://{_settings.SANITY_PROJECT_ID}.api.sanity.io"
            f"/v{_settings.SANITY_API_VERSION}/assets/images/{_settings.SANITY_DATASET}"
            f"?filename=upload.{ext}"
        )
        async with httpx.AsyncClient(timeout=httpx.Timeout(60.0, connect=10.0)) as client:
            asset_res = await client.post(
                asset_url,
                content=image_bytes,
                headers={
                    "Authorization": f"Bearer {_settings.SANITY_TOKEN}",
                    "Content-Type": content_type,
                }
            )
        if asset_res.status_code not in (200, 201):
            raise HTTPException(status_code=500, detail=f"Sanity asset upload failed: {asset_res.text}")

        asset_data = asset_res.json()
        asset_ref = asset_data["document"]["_id"]

        return {"status": "success", "assetRef": asset_ref}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/document/{doc_id}/image", dependencies=[Depends(verify_admin_token)])
async def upload_document_image(doc_id: str, request: Request):
    """
    Uploads a raw image file to Sanity assets, then patches the document's image field.
    Expects raw binary body with Content-Type: image/*, e.g. image/jpeg or image/png.
    """
    from app.core.config import get_settings
    _settings = get_settings()
    content_type = request.headers.get("content-type", "image/jpeg")
    image_bytes = await request.body()

    if not image_bytes:
        raise HTTPException(status_code=400, detail="Empty image body.")

    # Determine file extension from content-type
    ext_map = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}
    ext = ext_map.get(content_type.split(";")[0].strip(), "jpg")

    try:
        # pyrefly: ignore [missing-import]
        import httpx
        asset_url = (
            f"https://{_settings.SANITY_PROJECT_ID}.api.sanity.io"
            f"/v{_settings.SANITY_API_VERSION}/assets/images/{_settings.SANITY_DATASET}"
            f"?filename=upload.{ext}"
        )
        async with httpx.AsyncClient(timeout=httpx.Timeout(60.0, connect=10.0)) as client:
            asset_res = await client.post(
                asset_url,
                content=image_bytes,
                headers={
                    "Authorization": f"Bearer {_settings.SANITY_TOKEN}",
                    "Content-Type": content_type,
                }
            )
        if asset_res.status_code not in (200, 201):
            raise HTTPException(status_code=500, detail=f"Sanity asset upload failed: {asset_res.text}")

        asset_data = asset_res.json()
        asset_ref = asset_data["document"]["_id"]

        # Patch the document's image field with the new asset reference
        mutations = [{"patch": {"id": doc_id, "set": {
            "image": {"_type": "image", "asset": {"_type": "reference", "_ref": asset_ref}}
        }}}]
        await sanity_service.mutate_sanity(mutations)

        return {"status": "success", "assetRef": asset_ref}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/media/check-affected/{doc_id}", dependencies=[Depends(verify_admin_token)])
async def check_affected_records(doc_id: str):
    """
    Checks if a media asset is an import source and returns count of seeded records.
    """
    try:
        doc = await sanity_service.query_sanity(f'*[_id == "{doc_id}"][0]')
        if not doc:
            return {"is_import_file": False, "affected_count": 0, "filename": ""}
            
        original_filename = doc.get("originalFilename")
        if not original_filename:
            return {"is_import_file": False, "affected_count": 0, "filename": ""}
            
        # Count documents matching importFileId or importFilename, excluding the asset itself
        query = f'count(*[(importFileId == "{doc_id}" || importFilename == "{original_filename}") && _id != "{doc_id}"])'
        count = await sanity_service.query_sanity(query) or 0
        
        return {
            "is_import_file": True,
            "filename": original_filename,
            "affected_count": count
        }
    except Exception as e:
        print(f"ERROR checking affected records: {e}")
        return {"is_import_file": False, "affected_count": 0, "filename": ""}


@router.delete("/document/{doc_id}", dependencies=[Depends(verify_admin_token)])
async def delete_document(doc_id: str, delete_seeded: bool = False):
    """
    Unified document deletion endpoint.
    Admin users cannot be deleted for safety. Access points can be deleted.
    Can cascade delete any records that were imported/seeded from this file.
    """
    import re
    _UUID_RE = re.compile(
        r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        re.IGNORECASE
    )

    try:
        is_security_record = False
        record = None
        if _UUID_RE.match(doc_id) or doc_id.isdigit() or doc_id.startswith(("GMAU", "GMAP", "GMSL")):
            # Supabase security record
            record = await security_service.get_security_record_by_id(doc_id)
            if record:
                is_security_record = True
            elif doc_id.startswith(("GMAU", "GMAP", "GMSL")):
                raise HTTPException(status_code=404, detail="Security record not found in Supabase.")
            
        if is_security_record:
            table = record.get("_table")
            if table == "admin_users":
                raise HTTPException(status_code=400, detail="Security admin users cannot be deleted. Use Toggle to suspend.")
            elif table == "access_points":
                await security_service.delete_security_record(doc_id)
                return {"status": "success", "id": doc_id}
            else:
                raise HTTPException(status_code=400, detail=f"Deletion not supported for security table: {table}")
        else:
            base_id = doc_id[7:] if doc_id.startswith("drafts.") else doc_id
            mutations = [
                {"delete": {"id": base_id}},
                {"delete": {"id": f"drafts.{base_id}"}}
            ]
            if delete_seeded:
                doc = await sanity_service.query_sanity(f'*[_id == "{base_id}"][0]')
                if not doc:
                    doc = await sanity_service.query_sanity(f'*[_id == "drafts.{base_id}"][0]')
                if doc:
                    original_filename = doc.get("originalFilename")
                    query = f'*[(importFileId == "{base_id}" || importFileId == "drafts.{base_id}" || importFilename == "{original_filename}") && _id != "{base_id}" && _id != "drafts.{base_id}"][0...1000]._id'
                    seeded_ids = await sanity_service.query_sanity(query) or []
                    for sid in seeded_ids:
                        mutations.append({"delete": {"id": sid}})
                        
            result = await sanity_service.mutate_sanity(mutations)
            return {"status": "success", "id": base_id, "deleted_seeded_count": max(0, len(mutations) - 2), "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/document/{doc_id}/publish", dependencies=[Depends(verify_admin_token)])
async def publish_document(doc_id: str):
    """
    Publishes draft changes of a document to its published counterpart,
    and then deletes the draft.
    """
    base_id = doc_id[7:] if doc_id.startswith("drafts.") else doc_id
    draft_id = f"drafts.{base_id}"
    
    try:
        # Fetch the draft
        draft_doc = await sanity_service.query_sanity(f'*[_id == "{draft_id}"][0]')
        if not draft_doc:
            raise HTTPException(status_code=404, detail="No draft found to publish.")
            
        pub_doc = dict(draft_doc)
        pub_doc["_id"] = base_id
        for key in ("_createdAt", "_updatedAt", "_rev"):
            pub_doc.pop(key, None)
            
        mutations = [
            {"createOrReplace": pub_doc},
            {"delete": {"id": draft_id}}
        ]
        result = await sanity_service.mutate_sanity(mutations)
        return {"status": "success", "id": base_id, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/document/{doc_id}/discard", dependencies=[Depends(verify_admin_token)])
async def discard_document_draft(doc_id: str):
    """
    Discards the draft version of a document.
    If the document has never been published, discarding the draft deletes it completely.
    """
    base_id = doc_id[7:] if doc_id.startswith("drafts.") else doc_id
    draft_id = f"drafts.{base_id}"
    
    try:
        # Check if published version exists
        pub_exists = await sanity_service.query_sanity(f'count(*[_id == "{base_id}"])')
        
        # Delete the draft
        mutations = [{"delete": {"id": draft_id}}]
        result = await sanity_service.mutate_sanity(mutations)
        
        return {
            "status": "success",
            "id": base_id,
            "deleted_completely": pub_exists == 0,
            "result": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== BULK ACTIONS ==========

class BulkDeletePayload(BaseModel):
    ids: list

@router.delete("/bulk", dependencies=[Depends(verify_admin_token)])
async def bulk_delete_documents(payload: BulkDeletePayload):
    """Deletes multiple documents in a single Sanity mutation batch."""
    if not payload.ids:
        raise HTTPException(status_code=400, detail="No IDs provided.")
    try:
        mutations = [{"delete": {"id": doc_id}} for doc_id in payload.ids]
        result = await sanity_service.mutate_sanity(mutations)
        return {"status": "success", "deleted": len(payload.ids)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _auto_slugify(doc_type: str, data: dict):
    # Check if there is a 'slug' field in the schema
    schemas = schema_loader.get_schemas()
    schema = next((s for s in schemas if s["name"] == doc_type), None)
    if not schema:
        return None
    
    # Check if this schema has a slug field
    slug_field = next((f for f in schema["fields"] if f["name"] == "slug"), None)
    if not slug_field:
        return None
        
    # Find a source field
    source_field_name = slug_field.get("source")
    source_val = None
    if source_field_name:
        source_val = data.get(source_field_name)
    if not source_val:
        source_val = data.get("name") or data.get("title")
        
    if not source_val:
        return None
        
    slug_val = re.sub(r'[^a-z0-9\-]', '', str(source_val).lower().replace(' ', '-'))
    return {"_type": "slug", "current": slug_val}
async def _resolve_or_create_category(
    v: str,
    row_data: dict = None,
    category_cache: dict = None,
    import_file_id: str = None,
    import_filename: str = None,
    dry_run: bool = False
):
    try:
        v = str(v).strip()
        if not v:
            return None

        # Retrieve from cache (since it is guaranteed to be primed upfront)
        if category_cache:
            target_info = category_cache.get(v.lower())
            if target_info:
                target_id, _ = target_info
                return {"_type": "reference", "_ref": target_id}

        target_id = f"GMCAT-DRY-{v.lower().replace(' ', '-')}"
        return {"_type": "reference", "_ref": target_id}
    except Exception as e:
        print(f"ERROR: Category resolution failed: {e}")
        return None

async def _coerce_field(
    collection: str, 
    key: str, 
    value: str, 
    row_data: dict = None, 
    category_cache: dict = None,
    import_file_id: str = None,
    import_filename: str = None,
    split_categories: bool = False,
    dry_run: bool = False
):
    """Best-effort type coercion for imported field values based on dynamic schemas."""
    if value is None or str(value).strip() == '' or str(value).lower() == 'nan':
        return None
    v = str(value).strip()
    
    schemas = schema_loader.get_schemas()
    schema = next((s for s in schemas if s["name"] == collection), None)
    if not schema:
        if key in ('price', 'priceOverride'):
            try: return float(v)
            except: return None
        if key == 'availability':
            return v.lower() not in ('false', '0', 'no', 'out of stock')
        if key == 'keywords':
            return [k.strip() for k in v.split(',') if k.strip()]
        return v
        
    import re
    field = next((f for f in schema["fields"] if re.sub(r'[^a-z0-9]', '', f["name"].lower()) == key), None)
    if not field:
        return v
        
    ftype = field["type"]
    
    if ftype == "number":
        try: return float(v)
        except: return None
    elif ftype == "boolean":
        return v.lower() not in ('false', '0', 'no', 'out of stock')
    elif ftype == "array":
        return [k.strip() for k in v.split(',') if k.strip()]
    elif ftype == "reference":
        to_val = field.get("to")
        if isinstance(to_val, list):
            target_type = to_val[0]["type"] if isinstance(to_val[0], dict) else to_val[0]
        elif isinstance(to_val, str):
            target_type = to_val
        else:
            target_type = "category"
            
        if target_type == "category":
            if split_categories and v:
                import re
                parts = [p.strip() for p in re.split(r'[,/]', str(v)) if p.strip()]
                if len(parts) > 1:
                    first_ref = None
                    for idx, part in enumerate(parts):
                        part_ref = await _resolve_or_create_category(
                            part,
                            row_data,
                            category_cache,
                            import_file_id,
                            import_filename,
                            dry_run
                        )
                        if idx == 0:
                            first_ref = part_ref
                    return first_ref

            return await _resolve_or_create_category(
                v,
                row_data,
                category_cache,
                import_file_id,
                import_filename,
                dry_run
            )

        # Non-category references
        query = f'*[_type == "{target_type}" && (name == "{v}" || title == "{v}")][0]'
        try:
            target_doc = await sanity_service.query_sanity(query)
            if target_doc:
                return {"_type": "reference", "_ref": target_doc.get('_id')}
        except Exception as e:
            print(f"ERROR: Non-category reference lookup failed: {e}")
        return None
        
    return v


@router.get("/media", dependencies=[Depends(verify_admin_token)])
async def get_media_assets():
    """Fetches uploaded file and image assets from Sanity."""
    query = '*[_type in ["sanity.imageAsset", "sanity.fileAsset"]] | order(_createdAt desc)[0...100]'
    try:
        assets = await sanity_service.query_sanity(query)
        return {
            "success": True,
            "data": assets,
            "project_id": sanity_service.project_id,
            "dataset": sanity_service.dataset
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bulk/preview", dependencies=[Depends(verify_admin_token)])
async def preview_bulk_file(file: UploadFile = File(...)):
    """Parses a file and returns the first 5 rows and total count for frontend mapping preview."""
    content = await file.read()
    filename = (file.filename or '').lower()
    rows = []

    try:
        if filename.endswith('.json'):
            import json
            parsed = json.loads(content)
            rows = parsed if isinstance(parsed, list) else [parsed]
        elif filename.endswith('.csv'):
            import csv, io
            reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
            rows = [dict(r) for r in reader]
        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    # Check for category separators in any columns containing "category"
    category_separators_found = []
    if rows:
        import re
        category_keys = []
        for k in rows[0].keys():
            if k and re.search(r'category', str(k), re.IGNORECASE):
                category_keys.append(k)
        
        for row in rows:
            for k in category_keys:
                val = row.get(k)
                if val and isinstance(val, str) and (',' in val or '/' in val):
                    parts = [p.strip() for p in re.split(r'[,/]', val) if p.strip()]
                    if len(parts) > 1:
                        category_separators_found.append({
                            "column": k,
                            "original": val,
                            "split_preview": parts
                        })
                        if len(category_separators_found) >= 5:
                            break
            if len(category_separators_found) >= 5:
                break

    return {
        "success": True,
        "total": len(rows),
        "preview": rows[:5],
        "category_separators": category_separators_found
    }


def _generate_deterministic_id(collection: str, doc_payload: dict, fallback_seq: int, prefix: str) -> str:
    if collection == 'product':
        brand = str(doc_payload.get('brandName') or doc_payload.get('brandname') or '').strip()
        generic = str(doc_payload.get('genericName') or doc_payload.get('genericname') or '').strip()
        strength = str(doc_payload.get('strength') or '').strip()
        if brand or generic or strength:
            raw = f"prod-{brand}-{generic}-{strength}".lower()
            import re
            cleaned = re.sub(r'[^a-z0-9\-]', '', raw.replace(' ', '-').replace('/', '-').replace('+', '-').replace('(', '-').replace(')', '-'))
            cleaned = re.sub(r'-+', '-', cleaned).strip('-')
            return f"GM-{cleaned}"
    elif collection == 'category':
        cat = str(doc_payload.get('category') or '').strip()
        if cat:
            import re
            cleaned = re.sub(r'[^a-z0-9\-]', '', cat.lower().replace(' ', '-'))
            return f"GMCAT-{cleaned}"
    return f"{prefix}{str(fallback_seq).zfill(4)}"


@router.post("/bulk/import/{collection}", dependencies=[Depends(verify_admin_token)])
async def bulk_import_documents(
    collection: str, 
    file: UploadFile = File(...), 
    mappings: str = Form(None), 
    overwrite: bool = Form(False),
    duplicate_action: str = Form(None),
    id_columns: str = Form(None),
    dry_run: bool = Form(False),
    split_categories: bool = Form(False)
):
    """
    Parses CSV, XLSX, or JSON file and bulk-creates documents in the given collection.
    Columns/keys are matched to schema field names (case-insensitive or manual mapping).
    Supports dry run mode to analyze duplicates, custom ID generation columns, and real-time streaming progress.
    """
    from fastapi.responses import StreamingResponse
    import json

    if collection not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported collection: {collection}")

    content = await file.read()
    filename = (file.filename or '').lower()
    rows = []

    try:
        if filename.endswith('.json'):
            import json
            parsed = json.loads(content)
            rows = parsed if isinstance(parsed, list) else [parsed]

        elif filename.endswith('.csv'):
            import csv, io
            reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
            rows = [dict(r) for r in reader]

        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use CSV, XLSX, or JSON.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or contains no data rows.")

    # Check if a file with the same name already exists in Sanity
    if not overwrite:
        existing_assets = await sanity_service.query_sanity(
            f'*[_type in ["sanity.imageAsset", "sanity.fileAsset"] && originalFilename == "{file.filename}"]'
        )
        if existing_assets:
            raise HTTPException(
                status_code=409, 
                detail=f"file_exists:{file.filename}"
            )
    else:
        # Delete existing assets of same name AND all seeded/imported documents under it to keep it clean!
        existing_assets = await sanity_service.query_sanity(
            f'*[_type in ["sanity.imageAsset", "sanity.fileAsset"] && originalFilename == "{file.filename}"]'
        )
        if existing_assets:
            delete_mutations = []
            for asset in existing_assets:
                asset_id = asset["_id"]
                delete_mutations.append({"delete": {"id": asset_id}})
                # Query seeded documents to delete (like products and categories)
                query = f'*[(importFileId == "{asset_id}" || importFilename == "{file.filename}") && _id != "{asset_id}"][0...2000]._id'
                seeded_ids = await sanity_service.query_sanity(query) or []
                for sid in seeded_ids:
                    delete_mutations.append({"delete": {"id": sid}})
            
            if delete_mutations:
                try:
                    await sanity_service.mutate_sanity(delete_mutations)
                    print(f"DEBUG: Cascade-deleted existing asset '{file.filename}' and {len(delete_mutations)-len(existing_assets)} seeded documents for overwrite.")
                except Exception as del_err:
                    print(f"WARNING: Failed to cascade-delete duplicate assets: {del_err}")

    analysis_logs = []

    # Store file in assets database first to track import history (skip during dry_run to keep assets clean)
    asset_id = None
    if not dry_run:
        try:
            filename = file.filename or "bulk-import-file"
            asset_type = "images" if filename.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')) else "files"
            content_type = file.content_type or ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.endswith('.xlsx') else "text/csv")
            asset_doc = await sanity_service.upload_asset(
                asset_type=asset_type,
                file_bytes=content,
                filename=file.filename or "bulk-import-file",
                content_type=content_type
            )
            if asset_doc:
                asset_id = asset_doc.get('_id')
            log_msg = f"Successfully stored bulk import file '{file.filename}' in Sanity database (ID: {asset_id})"
            print(f"DEBUG: {log_msg}")
            analysis_logs.append(log_msg)
        except Exception as upload_err:
            log_msg = f"WARNING: Failed to store import file in Sanity assets: {upload_err}"
            print(log_msg)
            analysis_logs.append(log_msg)

    # Parse custom mappings if provided
    mappings_dict = None
    if mappings:
        try:
            import json
            mappings_dict = json.loads(mappings)
        except Exception as parse_err:
            print(f"ERROR: Failed to parse custom mappings JSON: {parse_err}")

    # Parse custom ID columns if provided
    id_cols_list = None
    if id_columns:
        try:
            import json
            id_cols_list = json.loads(id_columns)
        except Exception as id_parse_err:
            print(f"WARNING: Failed to parse id_columns JSON: {id_parse_err}")

    # 1. Build category cache upfront (batch creation and update to prevent sequential querying)
    category_cache = {}
    if collection == "product" and rows:
        try:
            import re
            import uuid
            # Identify columns mapping to 'category' and 'subcategory'
            category_cols = []
            subcategory_cols = []
            sample_row = rows[0]
            for col_name in sample_row.keys():
                target_field = col_name
                if mappings_dict:
                    target_field = mappings_dict.get(col_name)
                if target_field:
                    norm_field = re.sub(r'[^a-z0-9]', '', str(target_field).lower())
                    if norm_field in ('category', 'categories'):
                        category_cols.append(col_name)
                    elif norm_field in ('subcategory', 'subcategories'):
                        subcategory_cols.append(col_name)

            # Gather all category terms and their associated subcategories from the spreadsheet
            category_to_subs = {}
            for row in rows:
                sub_val = None
                for sub_col in subcategory_cols:
                    sv = row.get(sub_col)
                    if sv is not None and str(sv).strip() != '' and str(sv).lower() != 'nan':
                        sub_val = str(sv).strip()
                        break
                
                for cat_col in category_cols:
                    cv = row.get(cat_col)
                    if cv is not None and str(cv).strip() != '' and str(cv).lower() != 'nan':
                        val_str = str(cv).strip()
                        if split_categories:
                            parts = [p.strip() for p in re.split(r'[,/]', val_str) if p.strip()]
                        else:
                            parts = [val_str]
                        
                        for part in parts:
                            part_lower = part.lower()
                            if part_lower not in category_to_subs:
                                category_to_subs[part_lower] = {"title": part, "subcategories": set()}
                            if sub_val:
                                category_to_subs[part_lower]["subcategories"].add(sub_val)

            # Query existing categories from Sanity
            all_categories = await sanity_service.query_sanity('*[_type == "category"]')
            for cat in all_categories:
                cat_id = cat.get('_id')
                cat_name = cat.get('category')
                cat_title = cat.get('title')
                cat_cid = cat.get('categoryId')
                cat_subs = list(cat.get('subcategory') or [])
                for term in (cat_name, cat_title, cat_cid):
                    if term:
                        category_cache[str(term).strip().lower()] = (cat_id, cat_subs)

            # Determine upfront category mutations
            new_cat_mutations = []
            patch_cat_mutations = []
            
            for term_lower, info in category_to_subs.items():
                term_title = info["title"]
                sheet_subs = info["subcategories"]
                
                if term_lower in category_cache:
                    cat_id, existing_subs = category_cache[term_lower]
                    new_subs = [s for s in sheet_subs if s not in existing_subs]
                    if new_subs:
                        updated_subs = list(existing_subs) + new_subs
                        category_cache[term_lower] = (cat_id, updated_subs)
                        if not dry_run:
                            patch_cat_mutations.append({
                                "patch": {
                                    "id": cat_id,
                                    "set": {"subcategory": updated_subs}
                                }
                            })
                else:
                    mock_id = f"GMCAT-DRY-{term_lower.replace(' ', '-')}"
                    if not dry_run:
                        cat_id = f"GMCAT-{uuid.uuid4().hex[:4].upper()}"
                        new_doc = {
                            "_id": cat_id,
                            "_type": "category",
                            "categoryId": cat_id,
                            "category": term_title,
                            "title": term_title,
                            "name": term_title,
                            "subcategory": list(sheet_subs)
                        }
                        if asset_id:
                            new_doc["importFileId"] = asset_id
                        filename = file.filename or "bulk-import-file"
                        new_doc["importFilename"] = filename
                        
                        new_cat_mutations.append({"create": new_doc})
                        category_cache[term_lower] = (cat_id, list(sheet_subs))
                    else:
                        category_cache[term_lower] = (mock_id, list(sheet_subs))

            # Commit category mutations upfront
            if not dry_run:
                if new_cat_mutations:
                    await sanity_service.mutate_sanity(new_cat_mutations)
                    log_msg = f"Created {len(new_cat_mutations)} new categories upfront."
                    print(f"DEBUG: {log_msg}")
                    analysis_logs.append(log_msg)
                if patch_cat_mutations:
                    await sanity_service.mutate_sanity(patch_cat_mutations)
                    log_msg = f"Updated subcategories for {len(patch_cat_mutations)} existing categories upfront."
                    print(f"DEBUG: {log_msg}")
                    analysis_logs.append(log_msg)
            
            log_msg = f"Category cache primed and optimized with {len(category_cache)} entries."
            print(f"DEBUG: {log_msg}")
            analysis_logs.append(log_msg)
        except Exception as cache_err:
            log_msg = f"WARNING: Failed to batch-prime category cache: {cache_err}"
            print(log_msg)
            analysis_logs.append(log_msg)

    # 2. Build payload documents and pre-determine their IDs
    payloads = []
    generated_ids = []
    compiled_ids = set()
    
    prefix = PREFIX_MAP.get(collection, 'GM-')
    count_query = f'count(*[_type == "{collection}"])'
    try:
        existing_count = await sanity_service.query_sanity(count_query) or 0
    except:
        existing_count = 0

    for idx, row in enumerate(rows):
        try:
            import re
            normalized_row = {}
            for raw_k, raw_v in row.items():
                norm_k = re.sub(r'[^a-z0-9]', '', str(raw_k).lower())
                normalized_row[norm_k] = raw_v
                if mappings_dict:
                    target_f = mappings_dict.get(raw_k)
                    if target_f and target_f != "[Ignore]":
                        norm_k_custom = re.sub(r'[^a-z0-9]', '', target_f.lower())
                        normalized_row[norm_k_custom] = raw_v

            doc_payload = {"_type": collection}

            if mappings_dict:
                for k, v in row.items():
                    target_field = mappings_dict.get(k)
                    if target_field and target_field != "[Ignore]":
                        normalized_key = re.sub(r'[^a-z0-9]', '', target_field.lower())
                        coerced = await _coerce_field(
                            collection, 
                            normalized_key, 
                            v, 
                            row_data=normalized_row, 
                            category_cache=category_cache,
                            import_file_id=asset_id,
                            import_filename=file.filename,
                            split_categories=split_categories,
                            dry_run=dry_run
                        )
                        if coerced is not None:
                            doc_payload[target_field] = coerced
            else:
                for k, v in row.items():
                    normalized_key = re.sub(r'[^a-z0-9]', '', str(k).lower())
                    if normalized_key in ('id', 'type', 'rev', 'createdat', 'updatedat'):
                        continue
                    coerced = await _coerce_field(
                        collection, 
                        normalized_key, 
                        v, 
                        row_data=normalized_row, 
                        category_cache=category_cache,
                        import_file_id=asset_id,
                        import_filename=file.filename,
                        split_categories=split_categories,
                        dry_run=dry_run
                    )
                    if coerced is not None:
                        exact_key = normalized_key
                        schemas = schema_loader.get_schemas()
                        schema = next((s for s in schemas if s["name"] == collection), None)
                        if schema:
                            field = next((f for f in schema["fields"] if re.sub(r'[^a-z0-9]', '', f["name"].lower()) == normalized_key), None)
                            if field:
                                exact_key = field["name"]
                        doc_payload[exact_key] = coerced

            if 'name' not in doc_payload:
                brand = doc_payload.get('brandname') or doc_payload.get('brandName')
                generic = doc_payload.get('genericname') or doc_payload.get('genericName')
                if brand and generic:
                    doc_payload['name'] = f"{brand} ({generic})"
                elif brand:
                    doc_payload['name'] = brand
                elif generic:
                    doc_payload['name'] = generic

            slug_obj = _auto_slugify(collection, doc_payload)
            if slug_obj:
                doc_payload['slug'] = slug_obj

            # Generate Document ID
            if id_cols_list:
                # Custom deterministic ID generation based on selected columns from normalized spreadsheet row
                slug_parts = []
                for col in id_cols_list:
                    # Look up by normalized key first in the spreadsheet row data
                    norm_col = re.sub(r'[^a-z0-9]', '', col.lower())
                    val = normalized_row.get(norm_col)
                    if val is None:
                        # Fallback to key matching in doc_payload
                        val = doc_payload.get(col)
                        if isinstance(val, dict) and "_ref" in val:
                            val = val["_ref"].replace("GMCAT-DRY-", "")
                    
                    if val is not None and str(val).strip() != '' and str(val).lower() != 'nan':
                        slug_parts.append(str(val).strip().lower())
                
                if slug_parts:
                    combined = "-".join(slug_parts)
                    slug = re.sub(r'[^a-z0-9\-]', '-', combined)
                    slug = re.sub(r'-+', '-', slug).strip('-')
                    doc_id = f"{prefix}{collection}-{slug}"
                else:
                    seq = existing_count + idx + 1
                    doc_id = _generate_deterministic_id(collection, doc_payload, seq, prefix)
            else:
                seq = existing_count + idx + 1
                doc_id = _generate_deterministic_id(collection, doc_payload, seq, prefix)

            doc_payload["_id"] = doc_id

            # Skip duplicate row within the same file (O(1) lookups)
            if doc_id in compiled_ids:
                log_msg = f"Skipping duplicate row {idx+1} in spreadsheet (ID: {doc_id})"
                print(f"DEBUG: {log_msg}")
                analysis_logs.append(log_msg)
                continue
            compiled_ids.add(doc_id)

            if asset_id:
                doc_payload["importFileId"] = asset_id
            doc_payload["importFilename"] = file.filename or "bulk-import-file"

            payloads.append(doc_payload)
            generated_ids.append(doc_id)
        except Exception as payload_err:
            log_msg = f"WARNING: Skipping row {idx+1} in payload compilation: {payload_err}"
            print(log_msg)
            analysis_logs.append(log_msg)

    # 3. Check which IDs already exist in Sanity via high-speed query
    existing_ids = set()
    if generated_ids:
        try:
            import json
            # Chunk queries to avoid HTTP GET URL length limit (30 IDs per chunk)
            chunk_size = 30
            for k in range(0, len(generated_ids), chunk_size):
                sub_ids = generated_ids[k:k+chunk_size]
                query = f'*[_id in {json.dumps(sub_ids)}]._id'
                chunk_existing = await sanity_service.query_sanity(query)
                if chunk_existing:
                    existing_ids.update(chunk_existing)
            log_msg = f"Found {len(existing_ids)} duplicate IDs in database."
            print(f"DEBUG: {log_msg}")
            analysis_logs.append(log_msg)
        except Exception as query_err:
            log_msg = f"WARNING: Failed to query existing duplicate IDs: {query_err}"
            print(log_msg)
            analysis_logs.append(log_msg)

    # 4. Handle Dry Run Request
    if dry_run:
        return {
            "status": "dry_run",
            "total": len(payloads),
            "duplicates": len(existing_ids),
            "columns": list(rows[0].keys()) if rows else [],
            "logs": analysis_logs
        }

    # 5. Return Streaming Response for Live Progress Bar
    async def import_generator():
        import asyncio
        created, failed, errors = 0, 0, []
        BATCH_SIZE = 5
        payload_with_row = [(idx + 1, payload) for idx, payload in enumerate(payloads)]
        total = len(payload_with_row)

        async def execute_mutations(items, is_retry=False):
            nonlocal created, failed
            filtered_items = []
            for row_idx, payload in items:
                doc_id = payload["_id"]
                if doc_id in existing_ids:
                    if duplicate_action == "skip":
                        continue
                    elif duplicate_action == "overwrite":
                        filtered_items.append((row_idx, {"createOrReplace": payload}))
                    else:
                        filtered_items.append((row_idx, {"create": payload}))
                else:
                    filtered_items.append((row_idx, {"create": payload}))
                    
            if not filtered_items:
                return

            muts = [m for _, m in filtered_items]
            try:
                await sanity_service.mutate_sanity(muts)
                created += len(filtered_items)
            except Exception as batch_err:
                # If a batch mutation fails, retry each item sequentially to report correct row errors
                if not is_retry and len(filtered_items) > 1:
                    print(f"WARNING: Batch mutation failed, retrying sequentially: {batch_err}")
                    for row_idx, mut_op in filtered_items:
                        try:
                            await sanity_service.mutate_sanity([mut_op])
                            created += 1
                        except Exception as single_err:
                            failed += 1
                            errors.append({"row": row_idx, "error": str(single_err)})
                else:
                    for row_idx, _ in filtered_items:
                        failed += 1
                        errors.append({"row": row_idx, "error": str(batch_err)})

        processed = 0
        for i in range(0, len(payload_with_row), BATCH_SIZE):
            chunk = payload_with_row[i:i+BATCH_SIZE]
            
            # Execute mutations first for this batch
            err_before = len(errors)
            await execute_mutations(chunk)
            err_after = len(errors)
            
            # Extract row-level errors
            failed_rows = {err["row"]: err["error"] for err in errors[err_before:err_after]}
            
            # Yield progress reports with logs for each item in the chunk
            for row_idx, p in chunk:
                n = p.get('name') or p.get('brandName') or p.get('genericName') or p.get('title') or 'Record'
                doc_id = p["_id"]
                processed += 1
                
                # Check status
                if row_idx in failed_rows:
                    log_msg = f"Failed to import {n} (Row {row_idx}): {failed_rows[row_idx]}"
                    log_type = "error"
                elif doc_id in existing_ids and duplicate_action == "skip":
                    log_msg = f"Skipped duplicate {n} (Row {row_idx})"
                    log_type = "warning"
                elif doc_id in existing_ids and duplicate_action == "overwrite":
                    log_msg = f"Overwrote duplicate {n} (Row {row_idx})"
                    log_type = "info"
                else:
                    log_msg = f"Imported {n} (Row {row_idx})"
                    log_type = "info"
                
                progress_payload = {
                    "status": "progress",
                    "current_name": n,
                    "processed": processed,
                    "total": total,
                    "log": log_msg,
                    "log_type": log_type
                }
                yield json.dumps(progress_payload) + "\n"
                await asyncio.sleep(0.01)

        # Final progress update
        progress_payload = {
            "status": "progress",
            "current_name": "Wrapping up...",
            "processed": total,
            "total": total
        }
        yield json.dumps(progress_payload) + "\n"
        await asyncio.sleep(0.02)

        done_payload = {
            "status": "done",
            "imported": created,
            "failed": failed,
            "total": total,
            "errors": errors[:10]
        }
        yield json.dumps(done_payload) + "\n"

    return StreamingResponse(import_generator(), media_type="application/x-ndjson")


@router.get("/bulk/export/{collection}", dependencies=[Depends(verify_admin_token)])
async def export_documents(collection: str, fmt: str = "csv"):
    """Exports all documents from a collection as CSV or JSON."""
    if collection not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported collection: {collection}")
    try:
        projection = GROQ_PROJECTIONS.get(collection, '{ _id, _type, _createdAt }')
        groq_query = f'*[_type == "{collection}"] {projection} | order(_createdAt desc)'
        docs = await sanity_service.query_sanity(groq_query)
        if not docs:
            docs = []

        if fmt == 'json':
            import json
            # pyrefly: ignore [missing-import]
            from fastapi.responses import Response as FR
            return FR(content=json.dumps(docs, default=str, indent=2),
                      media_type='application/json',
                      headers={"Content-Disposition": f"attachment; filename={collection}.json"})
        else:
            import csv, io
            output = io.StringIO()
            if docs:
                keys = list(docs[0].keys())
                writer = csv.DictWriter(output, fieldnames=keys, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(docs)
            # pyrefly: ignore [missing-import]
            from fastapi.responses import Response as FR
            return FR(content=output.getvalue(),
                      media_type='text/csv',
                      headers={"Content-Disposition": f"attachment; filename={collection}.csv"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/security/users", dependencies=[Depends(verify_admin_token)])
async def create_security_user(request: Request):
    """
    Creates a new admin user in Supabase PostgreSQL (admin_users table).
    """
    try:
        data = await request.json()
        await security_service.create_admin_user(data)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/security/users", dependencies=[Depends(verify_admin_token)])
async def get_security_users():
    """
    Returns all admin users from Supabase PostgreSQL (admin_users table).
    """
    try:
        return await security_service.get_admin_users()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/security/logs", dependencies=[Depends(verify_admin_token)])
async def get_security_logs(page: int = 1, limit: int = 5):
    """
    Exposes a server-side paginated list of security audit logs.
    """
    try:
        results, total_count = await security_service.get_security_logs_paginated(page, limit)
        
        return {
            "status": "success",
            "logs": results,
            "total": total_count,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/security/access", dependencies=[Depends(verify_admin_token)])
async def get_security_access():
    """
    Returns all access point policies from Supabase PostgreSQL (access_points table).
    """
    try:
        return await security_service.get_access_points()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/security/user/status", dependencies=[Depends(verify_admin_token)])
async def update_user_status(payload: dict):
    """
    Toggles admin user active/suspended status in Supabase PostgreSQL (admin_users table).
    """
    username = payload.get("username")
    status = payload.get("status")
    if not username or not status:
        raise HTTPException(status_code=400, detail="Missing username or status in payload")
    try:
        await security_service.update_user_status(username, status)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/media", dependencies=[Depends(verify_admin_token)])
async def get_media_library():
    """
    Queries and returns all previously uploaded media assets (sanity.imageAsset) from Sanity.
    """
    try:
        query = '*[_type == "sanity.imageAsset"] { _id, url, originalFilename, _createdAt } | order(_createdAt desc)'
        results = await sanity_service.query_sanity(query)
        return {
            "status": "success",
            "media": results
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/security/role-permissions", dependencies=[Depends(verify_admin_token)])
async def get_role_permissions():
    """
    Returns the full RBAC permission matrix (v2 format):
      - roles: distinct role names from admin_users
      - permissions: { role: { collection: [actions] } }
    """
    try:
        roles = await security_service.get_distinct_roles()
        permissions = await security_service.get_role_permissions()
        return {"status": "success", "roles": roles, "permissions": permissions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/security/role-permissions", dependencies=[Depends(verify_admin_token)])
async def save_role_permissions(payload: dict):
    """
    Saves the full RBAC permission matrix atomically.
    Payload v2: { permissions: { role: { collection: [actions] } }, format: 'v2' }
    Payload legacy: { permissions: { collection: [roles] } }
    """
    permissions = payload.get("permissions")
    fmt = payload.get("format", "legacy")
    if not permissions or not isinstance(permissions, dict):
        raise HTTPException(status_code=400, detail="Missing or invalid permissions payload.")
    try:
        await security_service.save_role_permissions(permissions, fmt)
        await security_service.add_security_log(
            event="RBAC Action Control permissions matrix updated",
            user="System Admin",
            ip_address="127.0.0.1",
            status="SUCCESS",
            severity="WARNING"
        )
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))